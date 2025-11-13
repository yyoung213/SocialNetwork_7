import pandas as pd
from datetime import datetime
import subprocess
import sys

def save_to_excel(result_df, filename=None):
    """
    Save result_df DataFrame to Excel file with optimized formatting
    
    Args:
        result_df: pandas DataFrame containing the crawled data
        filename: Optional filename. If not provided, generates timestamp-based filename
    
    Returns:
        str: The filename of the saved Excel file
    """
    if result_df is None or result_df.empty:
        print("Error: result_df is None or empty. Cannot save to Excel.")
        return None
    
    # Generate filename
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"wanted_company_data_{timestamp}.xlsx"
    
    print(f"\nSaving DataFrame to Excel file: {filename}")
    print(f"DataFrame shape: {result_df.shape}")
    print(f"Columns: {list(result_df.columns)}")
    
    try:
        # Save to Excel
        result_df.to_excel(filename, index=False, engine='openpyxl')
        
        # Adjust column widths and row heights for better readability
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(filename)
        ws = wb.active
        
        # Set column widths
        # 기업명: 20, 나머지 텍스트 컬럼: 50
        column_widths = {
            '기업명': 20,
            '주요업무': 50,
            '자격요건': 50,
            '우대사항': 50
        }
        
        for col_idx, col_name in enumerate(result_df.columns, 1):
            col_letter = get_column_letter(col_idx)
            width = column_widths.get(col_name, 30)  # Default width 30
            ws.column_dimensions[col_letter].width = width
        
        # Enable text wrapping for all cells
        from openpyxl.styles import Alignment
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set row heights based on content length
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):  # Skip header row
            max_lines = 1
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    # Calculate approximate number of lines based on column width
                    col_letter = cell.column_letter
                    col_width = ws.column_dimensions[col_letter].width
                    # Estimate lines: text length / (column width * 1.2 for Korean characters)
                    estimated_lines = max(1, int(len(text) / (col_width * 1.2)) + 1)
                    max_lines = max(max_lines, estimated_lines)
                
                # Apply text wrapping
                cell.alignment = wrap_alignment
            
            # Set row height (minimum 15, approximately 15 points per line)
            row_height = max(15, min(15 * max_lines, 200))  # Cap at 200 points
            ws.row_dimensions[row_idx].height = row_height
        
        # Set header row height
        ws.row_dimensions[1].height = 20
        
        # Save the workbook
        wb.save(filename)
        print(f"✓ Successfully saved to {filename}")
        print(f"  Column widths adjusted for readability")
        print(f"  Row heights adjusted based on content length")
        return filename
    except Exception as e:
        print(f"✗ Error saving to Excel: {e}")
        print("Trying to install openpyxl...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"], 
                                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            result_df.to_excel(filename, index=False, engine='openpyxl')
            print(f"✓ Successfully saved to {filename}")
            return filename
        except Exception as e2:
            print(f"✗ Failed to save Excel file: {e2}")
            print("Saving as CSV instead...")
            csv_filename = filename.replace('.xlsx', '.csv')
            result_df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
            print(f"✓ Saved as CSV: {csv_filename}")
            return csv_filename

def main():
    """
    Example usage: Load result_df from WantedCrawling and save to Excel
    This function demonstrates how to use ExcelExtraction with WantedCrawling
    """
    print("=" * 80)
    print("Excel Extraction Module")
    print("=" * 80)
    print("\nThis module provides functionality to save result_df to Excel files.")
    print("\nUsage example:")
    print("  from WantedCrawling import main as crawl_main")
    print("  from ExcelExtraction import save_to_excel")
    print("  ")
    print("  result_df = crawl_main()")
    print("  if result_df is not None:")
    print("      save_to_excel(result_df)")

if __name__ == "__main__":
    main()

